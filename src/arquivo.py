import pandas as pd
from sklearn.linear_model import LinearRegression
from dateutil.relativedelta import relativedelta
import numpy as np
import openpyxl
import sys # Usado para parar o script em caso de erro
import statsmodels.api as sm  # Nova biblioteca para o SARIMA

# --- 1. Configurações Iniciais ---

# Nome do seu arquivo Excel
nome_arquivo = 'C:\Users\Livia Bontempo\OneDrive\amaral\Dados teste.xlsx'
# Nome da aba com dados "longos" (que você criou)
aba_dados_limpos = 'dados_para_analise' 
# Nome da sua aba original (como na imagem)
aba_original = 'Plan1' 
# Ano que queremos prever
ano_para_prever = 2025


# --- 2. Carregar e Preparar os Dados ---

# Carrega os dados da aba correta
try:
    df = pd.read_excel(nome_arquivo, sheet_name=aba_dados_limpos)
except Exception as e:
    print(f"Erro ao ler a aba '{aba_dados_limpos}': {e}")
    print("Verifique se o nome do arquivo e da aba estão corretos.")
    sys.exit() # Para o script

# Normaliza nomes de coluna e exibe para depuração
df.columns = df.columns.astype(str).str.strip()
print(f"Colunas detectadas na aba '{aba_dados_limpos}': {list(df.columns)}")

# Detecta automaticamente a coluna de datas 
if 'Data' in df.columns:
    date_col = 'Data'
else:
    for c in df.columns:
        cl = c.lower()
        if ('data' in cl) or ('mes' in cl) or ('mês' in cl) or ('month' in cl) or ('date' in cl):
            date_col = c
            break

if not date_col:
    print(f"Erro: não foi possível encontrar uma coluna de datas na aba '{aba_dados_limpos}'.")
    print("Colunas encontradas:", list(df.columns))
    print("Por favor renomeie a coluna de datas para 'Data' ou verifique a aba selecionada.")
    sys.exit()

# Tenta converter usando o formato esperado '%m/%Y' e, se falhar, usa um parser mais flexível
try:
    # tenta conversão estrita primeiro
    df['Data'] = pd.to_datetime(df[date_col], format='%m/%Y', errors='raise')
except Exception:
    # fallback: parser flexível (coerce para NaT em valores inválidos)
    df['Data'] = pd.to_datetime(df[date_col], dayfirst=True, errors='coerce')
    if df['Data'].isna().all():
        print(f"Erro: não foi possível converter a coluna '{date_col}' para datas.")
        print("Verifique os valores e o formato (ex: '08/2025' ou '08/25').")
        sys.exit()
    else:
        print(f"A coluna '{date_col}' foi convertida para datas usando um parser flexível. Alguns valores podem ter sido definidos como NaT.")

# Ordena os dados por data e remove dados faltantes
df = df.sort_values(by='Data')
df = df.dropna(subset=['QNT'])

# --- 3. Preparar os Dados para o Modelo de Série Temporal ---

# CRÍTICO: Modelos de série temporal precisam de um índice de data.
# Não precisamos mais do 'TimeStep'. A própria data é o 'X'.
df = df.set_index('Data')

# Nossa série 'y' é a coluna QNT
y = df['QNT']

# --- 4. Treinar o Modelo SARIMA ---

print("--- Treinando o Modelo SARIMA ---")
print("Isso pode levar alguns segundos...")

# Criamos o modelo
# order=(1, 1, 1) -> Parâmetros (p,d,q) para a TENDÊNCIA
# seasonal_order=(1, 1, 1, 12) -> Parâmetros (P,D,Q,s) para a SAZONALIDADE
# s=12 informa ao modelo que o padrão se repete a cada 12 meses.
try:
    model = sm.tsa.statespace.SARIMAX(
        y,
        order=(1, 1, 1),
        seasonal_order=(1, 1, 1, 12),
        enforce_stationarity=False,
        enforce_invertibility=False
    )
    
    # Treina o modelo
    results = model.fit(disp=False) # disp=False desliga os logs de treino
    
    print("--- Modelo Treinado com Sucesso ---")

except Exception as e:
    print(f"Erro ao treinar o modelo SARIMA: {e}")
    print("Verifique se há dados suficientes para a análise (pelo menos 2 ciclos sazonais, ex: 24 meses).")
    sys.exit()


# --- 5. Preparar Previsões Múltiplas ---

# Encontra a última data dos dados existentes
ultima_data = y.index.max()

# Descobre quantos meses precisamos prever
mes_final_desejado = 12 
ultimo_mes_dados = ultima_data.month
num_previsoes = mes_final_desejado - ultimo_mes_dados

if num_previsoes <= 0:
    print(f"\nOs dados já estão completos até Dezembro de {ano_para_prever}.")
    sys.exit()

print(f"\n--- Calculando {num_previsoes} Previsões para {ano_para_prever} ---")

# Faz a previsão
forecast_object = results.get_forecast(steps=num_previsoes)
# Pega as previsões (elas já vêm com as datas corretas)
previsoes_series = forecast_object.predicted_mean

# Lista para guardar nossas previsões
previsoes_finais = []

# Loop para processar as previsões
for proxima_data, previsao_qnt in previsoes_series.items():
    previsao_qnt_arredondada = round(previsao_qnt)
    
    previsoes_finais.append({
        'data': proxima_data,
        'mes_nome': proxima_data.strftime('%B'), # Nome do mês em Inglês
        'previsao': previsao_qnt_arredondada
    })
    
    print(f"Previsão para {proxima_data.strftime('%m/%Y')}: {previsao_qnt_arredondada}")

# --- 6. Escrever os Resultados na Aba "graficos" ---

# Define a aba e o local exato onde os dados serão escritos
aba_alvo = 'graficos'
coluna_alvo = 6  
linhas_alvo = [11, 12, 13, 14] 

print(f"\n--- Escrevendo previsões na planilha '{aba_alvo}' ---")

# Verifica se o número de previsões bate com o número de linhas
# (O script gerou 4 previsões (Set/Out/Nov/Dez) e você especificou 4 linhas)
if len(previsoes_finais) != len(linhas_alvo):
    print(f"Erro: O script gerou {len(previsoes_finais)} previsões, mas você especificou {len(linhas_alvo)} linhas.")
    print("Por favor, ajuste o número de meses a prever ou as 'linhas_alvo' no script.")
    sys.exit() # Para o script se houver divergência

try:
    # Carrega o arquivo Excel inteiro
    workbook = openpyxl.load_workbook(nome_arquivo)
    
    # Seleciona a aba "graficos"
    if aba_alvo in workbook.sheetnames:
        sheet = workbook[aba_alvo]
    else:
        # Se a aba não existir, cria uma nova
        print(f"Aviso: Aba '{aba_alvo}' não encontrada. Criando uma nova...")
        sheet = workbook.create_sheet(title=aba_alvo)

    # --- Loop para escrever cada previsão no local exato ---

    for i, previsao in enumerate(previsoes_finais):
        
        # Pega a linha da nossa lista (ex: linhas_alvo[0] = 11)
        linha_atual = linhas_alvo[i]
        
        # Pega o valor da previsão
        valor_previsao = previsao['previsao']
        
        # Pega o nome do mês (para o print de confirmação)
        mes_nome = previsao['mes_nome'] 
        
        # Escreve o valor na célula (linha_atual, coluna_alvo)
        # Ex: (linha=11, coluna=6) -> Célula F11
        sheet.cell(row=linha_atual, column=coluna_alvo).value = valor_previsao
        
        print(f"Valor '{valor_previsao}' (para {mes_nome}) salvo na célula F{linha_atual}")

    # Salva o arquivo UMA VEZ, após todas as mudanças
    workbook.save(nome_arquivo)
    
    print(f"\nSucesso! Todas as {len(previsoes_finais)} previsões foram salvas em '{nome_arquivo}'.")

except Exception as e:
    print(f"\nOcorreu um erro ao tentar escrever na planilha: {e}")
    print("As previsões NÃO foram salvas.")